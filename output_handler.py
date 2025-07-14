"""
Модул за обработка на изходни данни
Създава интерактивна карта, Excel файлове и чартове за анализ
"""

import folium
import pandas as pd
import requests
import json
from typing import List, Dict, Tuple, Optional
import os
import logging
from config import get_config, OutputConfig
from cvrp_solver import CVRPSolution, Route
from warehouse_manager import WarehouseAllocation
from input_handler import Customer
from osrm_client import get_distance_matrix_from_central_cache

# OpenPyXL imports за Excel стилове
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

logger = logging.getLogger(__name__)

# Настройки за различните типове превозни средства
VEHICLE_SETTINGS = {
    'internal_bus': {
        'color': 'blue',
        'icon': 'bus',
        'prefix': 'fa',
        'name': 'Вътрешен автобус'
    },
    'center_bus': {
        'color': 'red', 
        'icon': 'building',
        'prefix': 'fa',
        'name': 'Централен автобус'
    },
    'external_bus': {
        'color': 'red',
        'icon': 'truck',
        'prefix': 'fa', 
        'name': 'Външен автобус'
    }
}

# Цветове за всеки отделен автобус
BUS_COLORS = [
    '#FF0000',  # Червен
    '#00FF00',  # Зелен  
    '#0000FF',  # Син
    '#FFFF00',  # Жълт
    '#FF00FF',  # Магента
    '#00FFFF',  # Циан
    '#FFA500',  # Оранжев
    '#800080',  # Лилав
    '#008000',  # Тъмно зелен
    '#000080',  # Тъмно син
    '#800000',  # Бордо
    '#808000',  # Маслинен
    '#FF69B4',  # Розов
    '#32CD32',  # Лайм зелен
    '#8A2BE2',  # Синьо виолетов
    '#FF4500',  # Червено оранжев
    '#2E8B57',  # Морско зелен
    '#4682B4',  # Стоманено син
    '#D2691E',  # Шоколадов
    '#DC143C'   # Тъмно червен
]


class InteractiveMapGenerator:
    """Генератор на интерактивна карта"""
    
    def __init__(self, config: OutputConfig):
        self.config = config
        # Зареждаме централната матрица
        self.central_matrix = get_distance_matrix_from_central_cache([])
        self.use_osrm_routing = self.central_matrix is not None
        
        # Проверяваме дали OSRM е достъпен
        try:
            import requests
            from config import get_config
            osrm_config = get_config().osrm
            test_url = f"{osrm_config.base_url.rstrip('/')}/route/v1/driving/23.3,42.7;23.3,42.7"
            response = requests.get(test_url, timeout=5)
            if response.status_code == 200:
                logger.info("✅ OSRM сървър е достъпен - ще използвам реални маршрути")
                self.use_osrm_routing = True
            else:
                logger.warning("⚠️ OSRM сървър не отговаря - ще използвам прави линии")
                self.use_osrm_routing = False
        except Exception as e:
            logger.warning(f"⚠️ Не мога да се свържа с OSRM сървъра: {e}")
            logger.warning("   Ще използвам прави линии за маршрутите")
            self.use_osrm_routing = False
    
    def create_map(self, solution: CVRPSolution, warehouse_allocation: WarehouseAllocation,
                  depot_location: Tuple[float, float]) -> folium.Map:
        """Създава интерактивна карта с маршрутите"""
        logger.info("Създавам интерактивна карта")
        
        # Показваме OSRM статуса
        if self.use_osrm_routing:
            logger.info("🛣️ Използвам OSRM Route API за реални маршрути")
        else:
            logger.warning("📐 Използвам прави линии (OSRM недостъпен)")
        
        # Инициализация на картата
        route_map = folium.Map(
            location=depot_location,
            zoom_start=self.config.map_zoom_level,
            tiles='OpenStreetMap'
        )
        
        # Добавяне на депото
        self._add_depot_marker(route_map, depot_location)
        
        # Добавяне на център зоната
        from config import get_config
        center_location = get_config().locations.center_location
        center_zone_radius = get_config().locations.center_zone_radius_km
        if get_config().locations.enable_center_zone_priority:
            self._add_center_zone_circle(route_map, center_location, center_zone_radius)
        
        # Добавяне на маршрутите с OSRM геометрия
        if self.config.show_route_colors:
            self._add_routes_to_map(route_map, solution.routes, depot_location)
        
        # Добавяне на легенда
        self._add_legend(route_map, solution.routes)
        
        return route_map
    
    def _add_depot_marker(self, route_map: folium.Map, depot_location: Tuple[float, float]):
        """Добавя маркер за депото"""
        folium.Marker(
            depot_location,
            popup="<b>Депо/Стартова точка</b>",
            tooltip="Депо",
            icon=folium.Icon(color='black', icon='home', prefix='fa')
        ).add_to(route_map)
    
    def _add_center_zone_circle(self, route_map: folium.Map, center_location: Tuple[float, float], radius_km: float):
        """Добавя кръг за център зоната"""
        folium.Circle(
            location=center_location,
            radius=radius_km * 1000,  # Конвертираме в метри
            color='red',
            fill=True,
            fillColor='red',
            fillOpacity=0.1,
            popup=f"<b>Център зона</b><br>Радиус: {radius_km} км",
            tooltip="Център зона"
        ).add_to(route_map)
        
        # Добавяме маркер за центъра
        folium.Marker(
            location=center_location,
            popup=f"<b>Център</b><br>Координати: {center_location[0]:.6f}, {center_location[1]:.6f}<br>Радиус зона: {radius_km} км",
            icon=folium.Icon(color='red', icon='star'),
            tooltip="Център"
        ).add_to(route_map)
    
    def _get_osrm_route_geometry(self, start_coords: Tuple[float, float],
                                end_coords: Tuple[float, float]) -> List[Tuple[float, float]]:
        """Получава реална геометрия на маршрута от OSRM Route API"""
        try:
            import requests
            from config import get_config
            
            # OSRM Route API заявка за пълна геометрия
            osrm_config = get_config().osrm
            base_url = osrm_config.base_url.rstrip('/')
            
            # Форматираме координатите за OSRM (lon,lat формат)
            start_lon, start_lat = start_coords[1], start_coords[0]
            end_lon, end_lat = end_coords[1], end_coords[0]
            
            route_url = f"{base_url}/route/v1/driving/{start_lon:.6f},{start_lat:.6f};{end_lon:.6f},{end_lat:.6f}?geometries=geojson&overview=full&steps=false"
            
            response = requests.get(route_url, timeout=10)
            response.raise_for_status()
            
            data = response.json()
            
            if data['code'] == 'Ok' and data['routes']:
                route = data['routes'][0]
                coordinates = route['geometry']['coordinates']
                
                # Конвертираме от [lon,lat] към [lat,lon] за Folium
                geometry = [(coord[1], coord[0]) for coord in coordinates]
                
                logger.debug(f"✅ OSRM геометрия получена: {len(geometry)} точки")
                return geometry
            else:
                logger.warning(f"OSRM Route API грешка: {data.get('message', 'Неизвестна грешка')}")
            return [start_coords, end_coords]
            
        except Exception as e:
            logger.warning(f"Грешка при OSRM Route API заявка: {e}")
            # Fallback към права линия
            return [start_coords, end_coords]
    
    def _get_full_route_geometry(self, waypoints: List[Tuple[float, float]]) -> List[Tuple[float, float]]:
        """Получава пълната геометрия за маршрут с множество точки от OSRM.
        Ако има твърде много точки, използваме fallback за по-бърза работа.
        """
        if len(waypoints) < 2:
            return waypoints

        # ОПТИМИЗАЦИЯ: Ако маршрутът има твърде много точки, не търсим пълна геометрия,
        # а чертаем сегменти, за да не претоварваме OSRM и да ускорим процеса.
        MAX_WAYPOINTS_FOR_FULL_GEOMETRY = 15
        if len(waypoints) > MAX_WAYPOINTS_FOR_FULL_GEOMETRY:
            logger.info(f"🌀 Маршрутът има {len(waypoints)} точки (> {MAX_WAYPOINTS_FOR_FULL_GEOMETRY}). "
                        f"Използвам опростена геометрия (сегменти) за по-бърза работа.")
            full_geometry = []
            for i in range(len(waypoints) - 1):
                # За всеки сегмент взимаме геометрията (или права линия при грешка)
                segment_geometry = self._get_osrm_route_geometry(waypoints[i], waypoints[i+1])
                if i > 0:
                    # Премахваме първата точка, за да няма дублиране
                    segment_geometry = segment_geometry[1:]
                full_geometry.extend(segment_geometry)
            return full_geometry

        try:
            import requests
            from config import get_config
            
            # OSRM Route API заявка за целия маршрут
            osrm_config = get_config().osrm
            base_url = osrm_config.base_url.rstrip('/')
            
            # Форматираме всички координати за OSRM (lon,lat формат)
            coords_str = ';'.join([f"{lon:.6f},{lat:.6f}" for lat, lon in waypoints])
            
            route_url = f"{base_url}/route/v1/driving/{coords_str}?geometries=geojson&overview=full&steps=false"
            
            response = requests.get(route_url, timeout=15)
            response.raise_for_status()
            
            data = response.json()
            
            if data['code'] == 'Ok' and data['routes']:
                route = data['routes'][0]
                coordinates = route['geometry']['coordinates']
                
                # Конвертираме от [lon,lat] към [lat,lon] за Folium
                geometry = [(coord[1], coord[0]) for coord in coordinates]
                
                logger.info(f"✅ OSRM маршрут геометрия получена: {len(geometry)} точки за {len(waypoints)} waypoints")
                return geometry
            else:
                logger.warning(f"OSRM Route API грешка за пълен маршрут: {data.get('message', 'Неизвестна грешка')}")
                return waypoints
                
        except Exception as e:
            logger.warning(f"Грешка при OSRM Route API заявка за пълен маршрут: {e}")
            # Fallback към последователност от прави линии
            full_geometry = []
            for i in range(len(waypoints) - 1):
                segment = self._get_osrm_route_geometry(waypoints[i], waypoints[i + 1])
                if i == 0:
                    full_geometry.extend(segment)
                else:
                    full_geometry.extend(segment[1:])  # Пропускаме дублираната точка
            
            return full_geometry if full_geometry else waypoints
    
    def _add_routes_to_map(self, route_map: folium.Map, routes: List[Route], depot_location: Tuple[float, float]):
        """Добавя маршрутите на картата с OSRM геометрия и филтър за бусовете"""
        # Създаваме FeatureGroup за всеки автобус
        bus_layers = {}
        
        for route_idx, route in enumerate(routes):
            vehicle_settings = VEHICLE_SETTINGS.get(route.vehicle_type.value, {
                'color': 'gray', 
                'icon': 'circle',
                'prefix': 'fa',
                'name': 'Неизвестен'
            })
            
            # Всеки автобус получава уникален цвят
            bus_color = BUS_COLORS[route_idx % len(BUS_COLORS)]
            bus_id = f"bus_{route_idx + 1}"
            
            # Създаваме FeatureGroup за този автобус
            bus_layer = folium.FeatureGroup(name=f"🚌 Автобус {route_idx + 1} ({len(route.customers)} клиента)")
            bus_layers[bus_id] = bus_layer
            
            # Добавяне на клиентските маркери с номерация
            for client_idx, customer in enumerate(route.customers):
                if customer.coordinates:
                    # Създаваме номериран маркер
                    client_number = client_idx + 1
                    
                    # HTML за номерирано пинче с уникален цвят на автобуса
                    icon_html = f'''
                    <div style="
                        background-color: {bus_color};
                        border: 3px solid white;
                        border-radius: 50%;
                        width: 30px;
                        height: 30px;
                        display: flex;
                        justify-content: center;
                        align-items: center;
                        font-weight: bold;
                        font-size: 14px;
                        color: white;
                        text-shadow: 1px 1px 1px rgba(0,0,0,0.7);
                    ">{client_number}</div>
                    '''
                    
                    popup_text = f"""
                    <div style="font-family: Arial, sans-serif;">
                        <h4 style="margin: 0; color: {bus_color};">
                            Автобус {route_idx + 1} - {vehicle_settings['name']}
                        </h4>
                        <hr style="margin: 5px 0;">
                        <b>Клиент:</b> {customer.name}<br>
                        <b>ID:</b> {customer.id}<br>
                        <b>Ред в маршрута:</b> #{client_number}<br>
                        <b>Обем:</b> {customer.volume:.2f} ст.<br>
                        <b>Координати:</b> {customer.coordinates[0]:.6f}, {customer.coordinates[1]:.6f}
                    </div>
                    """
                    
                    # Добавяме номерирания маркер в слоя на автобуса
                    marker = folium.Marker(
                        customer.coordinates,
                        popup=folium.Popup(popup_text, max_width=300),
                        tooltip=f"#{client_number}: {customer.name}",
                        icon=folium.DivIcon(
                            html=icon_html,
                            icon_size=(30, 30),
                            icon_anchor=(15, 15),
                            popup_anchor=(0, -15)
                        )
                    )
                    marker.add_to(bus_layer)
            
            # Създаваме пълния маршрут: депо -> клиенти -> депо
            if route.customers and self.use_osrm_routing:
                logger.info(f"🛣️ Получавам OSRM маршрут за Автобус {route_idx + 1} с {len(route.customers)} клиента")
                
                # Подготвяме всички waypoints
                waypoints = [depot_location]
                for customer in route.customers:
                    if customer.coordinates:
                        waypoints.append(customer.coordinates)
                waypoints.append(depot_location)  # Връщане в депото
                
                # Получаваме реалната геометрия от OSRM
                try:
                    route_geometry = self._get_full_route_geometry(waypoints)
                    
                    if len(route_geometry) > 2:
                        # Създаваме popup с информация за маршрута
                        popup_text = f"""
                        <div style="font-family: Arial, sans-serif;">
                            <h4 style="margin: 0; color: {bus_color};">
                                🚌 Автобус {route_idx + 1} - {vehicle_settings['name']}
                            </h4>
                            <hr style="margin: 5px 0;">
                            <b>OSRM маршрут:</b> ✅<br>
                            <b>Клиенти:</b> {len(route.customers)}<br>
                            <b>Разстояние:</b> {route.total_distance_km:.1f} км<br>
                            <b>Време:</b> {route.total_time_minutes:.0f} мин<br>
                            <b>Обем:</b> {route.total_volume:.1f} ст.<br>
                            <b>Геометрия:</b> {len(route_geometry)} точки
                        </div>
                        """
                        
                        # Създаваме линията в слоя на автобуса
                        polyline = folium.PolyLine(
                            route_geometry,
                            color=bus_color,
                            weight=4,
                            opacity=0.8,
                            popup=folium.Popup(popup_text, max_width=300)
                        )
                        polyline.add_to(bus_layer)
                        logger.info(f"✅ OSRM маршрут добавен за Автобус {route_idx + 1}: {len(route_geometry)} точки")
                    else:
                        # Fallback към прави линии
                        popup_text = f"""
                        <div style="font-family: Arial, sans-serif;">
                            <h4 style="margin: 0; color: {bus_color};">
                                🚌 Автобус {route_idx + 1} - {vehicle_settings['name']}
                            </h4>
                            <hr style="margin: 5px 0;">
                            <b>OSRM маршрут:</b> ⚠️ (прави линии)<br>
                            <b>Клиенти:</b> {len(route.customers)}<br>
                            <b>Разстояние:</b> {route.total_distance_km:.1f} км<br>
                            <b>Време:</b> {route.total_time_minutes:.0f} мин<br>
                            <b>Обем:</b> {route.total_volume:.1f} ст.
                        </div>
                        """
                        
                        polyline = folium.PolyLine(
                            waypoints,
                            color=bus_color,
                            weight=3,
                            opacity=0.6,
                            popup=folium.Popup(popup_text, max_width=300),
                            dashArray='5, 5'  # Пунктирана линия за показване че не е OSRM
                        )
                        polyline.add_to(bus_layer)
                        logger.warning(f"⚠️ Използвам прави линии за Автобус {route_idx + 1}")
                        
                except Exception as e:
                    logger.error(f"❌ Грешка при OSRM маршрут за Автобус {route_idx + 1}: {e}")
                    # Fallback към прави линии
                    waypoints = [depot_location]
                    for customer in route.customers:
                        if customer.coordinates:
                            waypoints.append(customer.coordinates)
                    waypoints.append(depot_location)
                    
                    popup_text = f"""
                    <div style="font-family: Arial, sans-serif;">
                        <h4 style="margin: 0; color: {bus_color};">
                            🚌 Автобус {route_idx + 1} - {vehicle_settings['name']}
                        </h4>
                        <hr style="margin: 5px 0;">
                        <b>OSRM маршрут:</b> ❌ (fallback)<br>
                        <b>Клиенти:</b> {len(route.customers)}<br>
                        <b>Разстояние:</b> {route.total_distance_km:.1f} км<br>
                        <b>Време:</b> {route.total_time_minutes:.0f} мин<br>
                        <b>Обем:</b> {route.total_volume:.1f} ст.
                    </div>
                    """
                    
                    polyline = folium.PolyLine(
                        waypoints,
                        color=bus_color,
                        weight=3,
                        opacity=0.6,
                        popup=folium.Popup(popup_text, max_width=300),
                        dashArray='5, 5'
                    )
                    polyline.add_to(bus_layer)
            
            elif route.customers:
                # Fallback към прави линии ако OSRM е изключен
                waypoints = [depot_location]
                for customer in route.customers:
                    if customer.coordinates:
                        waypoints.append(customer.coordinates)
                waypoints.append(depot_location)
                
                polyline = folium.PolyLine(
                    waypoints,
                    color=bus_color,
                    weight=3,
                    opacity=0.8,
                    popup=f"🚌 Автобус {route_idx + 1} - {vehicle_settings['name']}"
                )
                polyline.add_to(bus_layer)
        
        # Добавяме всички слоеве на автобусите към картата
        for bus_layer in bus_layers.values():
            bus_layer.add_to(route_map)
        
        # Добавяме LayerControl за филтър
        folium.LayerControl(
            position='topright',
            collapsed=False,
            overlay=True,
            control=True
                ).add_to(route_map)
    
    def _add_legend(self, route_map: folium.Map, routes: List[Route]):
        """Добавя легенда на картата с информация за маршрутите"""
        # Изчисляваме статистики
        total_distance = sum(route.total_distance_km for route in routes)
        total_time = sum(route.total_time_minutes for route in routes)
        total_volume = sum(route.total_volume for route in routes)
        osrm_routes = sum(1 for route in routes if self.use_osrm_routing)
        
        legend_html = f'''
        <div style="position: fixed; 
                    top: 10px; left: 10px; width: 280px; height: auto; 
                    background-color: white; border:2px solid grey; z-index:9999; 
                    font-size:14px; padding: 10px; border-radius: 5px;
                    box-shadow: 0 0 15px rgba(0,0,0,0.2);">
        <h4 style="margin-top:0; margin-bottom:10px; text-align: center;">🗺️ Информация</h4>
        '''
        
        # Добавяме депо
        legend_html += '''
        <p style="margin: 5px 0;">
            <i class="fa fa-home" style="color: black; margin-right: 8px;"></i>
            Депо
        </p>
        '''
        
        # Добавяме информация за филтъра
        legend_html += '''
        <hr style="margin: 10px 0;">
        <p style="margin: 5px 0; font-weight: bold;">🚌 Филтър на автобуси:</p>
        <p style="margin: 5px 0; font-size: 12px; color: #666;">
            Използвай контрола в горния десен ъгъл за показване/скриване на отделни автобуси
            </p>
            '''
        
        # Добавяме информация за OSRM маршрутите
        osrm_status = "🛣️ OSRM маршрути" if self.use_osrm_routing else "📐 Прави линии"
        
        legend_html += f'''
        <hr style="margin: 10px 0;">
        <p style="margin: 5px 0; font-size: 12px; color: #666;">
            Числата показват реда на посещение<br>
            {osrm_status}
        </p>
        '''
        
        # Добавяме статистики
        legend_html += f'''
        <hr style="margin: 10px 0;">
        <p style="margin: 5px 0; font-size: 12px; font-weight: bold;">
            📊 Статистики:
        </p>
        <p style="margin: 3px 0; font-size: 11px; color: #555;">
            • Общо разстояние: {total_distance:.1f} км<br>
            • Общо време: {total_time:.0f} мин<br>
            • Общ обем: {total_volume:.1f} ст.<br>
            • OSRM маршрути: {osrm_routes}/{len(routes)}
        </p>
        </div>
        '''
        
        # Добавяме легендата към картата
        legend_element = folium.Element(legend_html)
        route_map.get_root().add_child(legend_element)
    
    def save_map(self, route_map: folium.Map, file_path: Optional[str] = None) -> str:
        """Записва картата във файл"""
        file_path = file_path or self.config.map_output_file
        
        os.makedirs(os.path.dirname(file_path), exist_ok=True)
        route_map.save(file_path)
        
        logger.info(f"Интерактивна карта записана в {file_path}")
        return file_path


class ExcelExporter:
    """Експортър на Excel файлове"""
    
    def __init__(self, config: OutputConfig):
        self.config = config
    
    def export_all_to_single_excel(self, solution: CVRPSolution, warehouse_customers: List[Customer]) -> str:
        """Експортира всички данни в един Excel файл с отделни sheets"""
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        # Създаваме основния файл
        file_path = os.path.join(self.config.excel_output_dir, "cvrp_report.xlsx")
        os.makedirs(os.path.dirname(file_path), exist_ok=True)
        
        wb = Workbook()
        
        # Премахваме default sheet
        if wb.active:
            wb.remove(wb.active)
        
        # 1. SHEET: Маршрути (Vehicle Routes)
        if solution.routes:
            self._create_routes_sheet(wb, solution)
        
        # 2. SHEET: Необслужени клиенти (Unserved Customers)
        if warehouse_customers:
            self._create_unserved_sheet(wb, warehouse_customers)
        
        # 3. SHEET: Обобщение (Summary)
        self._create_summary_sheet(wb, solution, warehouse_customers)
        
        # 4. SHEET: Статистики по автобуси (Vehicle Statistics)
        if solution.routes:
            self._create_vehicle_stats_sheet(wb, solution)
        
        # Записваме файла
        wb.save(file_path)
        logger.info(f"Общ Excel отчет записан в {file_path}")
        return file_path
    
    def _create_routes_sheet(self, wb, solution: CVRPSolution):
        """Създава sheet с маршрутите"""
        ws = wb.create_sheet("Маршрути")
        
        # Заглавни редове
        headers = [
            'Маршрут', 'Превозно средство', 'Ред в маршрута', 
            'ID клиент', 'Име клиент', 'Обем (ст.)', 'GPS координати',
            'Разстояние до центъра (км)', 'Депо стартова точка',
            'Разстояние от предишен (км)', 'Накоплено разстояние (км)',
            'Време от предишен (мин)', 'Накоплено време (мин)'
        ]
        
        # Стилове за заглавния ред
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Добавяме заглавния ред
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Данни за маршрутите
        row = 2
        center_location = get_config().locations.center_location
        
        for i, route in enumerate(solution.routes):
            vehicle_name = VEHICLE_SETTINGS.get(route.vehicle_type.value, {}).get('name', 'Неизвестен')
            
            # Изчисляваме разстоянията и времената между клиентите
            cumulative_distance = 0
            cumulative_time = 0
            previous_customer_coords = route.depot_location  # Започваме от депото
            
            for j, customer in enumerate(route.customers):
                # Изчисляваме разстоянието до центъра
                distance_to_center = self._calculate_distance_to_center(customer.coordinates, center_location) if customer.coordinates else 0.0
                
                # Изчисляваме разстоянието от предишния клиент
                distance_from_previous = self._calculate_distance_between_points(
                    previous_customer_coords, customer.coordinates
                ) if customer.coordinates else 0.0
                cumulative_distance += distance_from_previous
                
                # Изчисляваме времето от предишния клиент (приблизително)
                time_from_previous = self._calculate_time_between_points(
                    previous_customer_coords, customer.coordinates
                ) if customer.coordinates else 0.0
                cumulative_time += time_from_previous
                
                # Проверяваме дали клиентът е в център зоната
                center_zone_radius = get_config().locations.center_zone_radius_km
                is_in_center_zone = distance_to_center <= center_zone_radius
                
                data = [
                    i + 1,  # Маршрут
                    vehicle_name,  # Превозно средство
                    j + 1,  # Ред в маршрута
                    customer.id,  # ID клиент
                    customer.name,  # Име клиент
                    customer.volume,  # Обем
                    customer.original_gps_data,  # GPS
                    round(distance_to_center, 2),  # Разстояние до центъра
                    f"{route.depot_location[0]:.6f}, {route.depot_location[1]:.6f}",  # Депо
                    round(distance_from_previous, 2),  # Разстояние от предишен
                    round(cumulative_distance, 2),  # Накоплено разстояние
                    round(time_from_previous, 1),  # Време от предишен
                    round(cumulative_time, 1),  # Накоплено време
                    "ДА" if is_in_center_zone else "НЕ"  # В център зоната
                ]
                
                for col, value in enumerate(data, 1):
                    ws.cell(row=row, column=col, value=value)
                
                row += 1
                previous_customer_coords = customer.coordinates
        
        # Автоматично разширяване на колоните
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _create_unserved_sheet(self, wb, warehouse_customers: List[Customer]):
        """Създава sheet с необслужените клиенти"""
        ws = wb.create_sheet("Необслужени клиенти")
        
        headers = [
            'ID', 'Име', 'Обем (ст.)', 'GPS координати', 
            'Latitude', 'Longitude', 'Разстояние до центъра (км)'
        ]
        
        # Стилове за заглавния ред
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="C5504B", end_color="C5504B", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Добавяме заглавния ред
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Данни
        center_location = get_config().locations.center_location
        row = 2
        
        for customer in warehouse_customers:
            distance_to_center = self._calculate_distance_to_center(customer.coordinates, center_location)
            
            data = [
                customer.id,
                customer.name,
                customer.volume,
                customer.original_gps_data,
                customer.coordinates[0] if customer.coordinates else '',
                customer.coordinates[1] if customer.coordinates else '',
                round(distance_to_center, 2)
            ]
            
            for col, value in enumerate(data, 1):
                ws.cell(row=row, column=col, value=value)
            
            row += 1
        
        # Автоматично разширяване на колоните
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _create_summary_sheet(self, wb, solution: CVRPSolution, warehouse_customers: List[Customer]):
        """Създава sheet с обобщение"""
        ws = wb.create_sheet("Обобщение")
        
        # Стилове
        title_font = Font(bold=True, size=14)
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        
        # Заглавие
        ws['A1'] = "CVRP ОТЧЕТ - ОБОБЩЕНИЕ"
        ws['A1'].font = title_font
        
        # Основни статистики
        row = 3
        stats = [
            ("Общо клиенти", len(solution.routes) + len(warehouse_customers)),
            ("Обслужени клиенти", sum(len(route.customers) for route in solution.routes)),
            ("Необслужени клиенти", len(warehouse_customers)),
            ("Брой маршрути", len(solution.routes)),
            ("Общо разстояние (км)", round(solution.total_distance_km, 2)),
            ("Общо време (мин)", round(solution.total_time_minutes, 2)),
            ("Общ обем (ст.)", round(sum(route.total_volume for route in solution.routes), 2))
        ]
        
        for stat_name, stat_value in stats:
            ws[f'A{row}'] = stat_name
            ws[f'A{row}'].font = header_font
            ws[f'B{row}'] = stat_value
            row += 1
        
        # Статистики по тип автобус
        row += 2
        ws[f'A{row}'] = "СТАТИСТИКИ ПО ТИП АВТОБУС"
        ws[f'A{row}'].font = title_font
        row += 1
        
        vehicle_stats = {}
        for route in solution.routes:
            vehicle_type = route.vehicle_type.value
            if vehicle_type not in vehicle_stats:
                vehicle_stats[vehicle_type] = {
                    'count': 0, 'distance': 0, 'volume': 0, 'customers': 0
                }
            vehicle_stats[vehicle_type]['count'] += 1
            vehicle_stats[vehicle_type]['distance'] += route.total_distance_km
            vehicle_stats[vehicle_type]['volume'] += route.total_volume
            vehicle_stats[vehicle_type]['customers'] += len(route.customers)
        
        # Заглавни редове за статистики
        headers = ['Тип автобус', 'Брой маршрути', 'Общо разстояние (км)', 'Общ обем (ст.)', 'Общо клиенти']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
        row += 1
        
        # Данни за статистики
        for vehicle_type, stats in vehicle_stats.items():
            vehicle_name = VEHICLE_SETTINGS.get(vehicle_type, {}).get('name', vehicle_type)
            data = [
                vehicle_name,
                stats['count'],
                round(stats['distance'], 2),
                round(stats['volume'], 2),
                stats['customers']
            ]
            for col, value in enumerate(data, 1):
                ws.cell(row=row, column=col, value=value)
            row += 1
        
        # Автоматично разширяване на колоните
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _create_vehicle_stats_sheet(self, wb, solution: CVRPSolution):
        """Създава sheet със статистики по отделни автобуси"""
        ws = wb.create_sheet("Статистики по автобуси")
        
        headers = [
            'Маршрут', 'Тип автобус', 'Брой клиенти', 'Общ обем (ст.)',
            'Разстояние (км)', 'Време (мин)', 'Капацитет използване (%)',
            'Средно разстояние до центъра (км)', 'Депо стартова точка'
        ]
        
        # Стилове за заглавния ред
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Добавяме заглавния ред
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Данни
        center_location = get_config().locations.center_location
        row = 2
        
        for i, route in enumerate(solution.routes):
            vehicle_name = VEHICLE_SETTINGS.get(route.vehicle_type.value, {}).get('name', 'Неизвестен')
            
            # Изчисляваме средното разстояние до центъра
            distances_to_center = []
            for customer in route.customers:
                distance = self._calculate_distance_to_center(customer.coordinates, center_location)
                distances_to_center.append(distance)
            avg_distance_to_center = sum(distances_to_center) / len(distances_to_center) if distances_to_center else 0
            
            # Капацитет използване (трябва да вземем capacity от config)
            vehicle_config = self._get_vehicle_config(route.vehicle_type)
            capacity_usage = 0
            if vehicle_config and vehicle_config.capacity > 0:
                capacity_usage = (route.total_volume / vehicle_config.capacity * 100)
            
            data = [
                i + 1,  # Маршрут
                vehicle_name,  # Тип автобус
                len(route.customers),  # Брой клиенти
                round(route.total_volume, 2),  # Общ обем
                round(route.total_distance_km, 2),  # Разстояние
                round(route.total_time_minutes, 2),  # Време
                round(capacity_usage, 1),  # Капацитет използване
                round(avg_distance_to_center, 2),  # Средно разстояние до центъра
                f"{route.depot_location[0]:.6f}, {route.depot_location[1]:.6f}"  # Депо
            ]
            
            for col, value in enumerate(data, 1):
                ws.cell(row=row, column=col, value=value)
            
            row += 1
        
        # Автоматично разширяване на колоните
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _calculate_distance_to_center(self, coordinates: Optional[Tuple[float, float]], center_location: Tuple[float, float]) -> float:
        """Изчислява разстоянието до центъра в км"""
        if not coordinates or not center_location:
            return 0.0
        
        from math import radians, sin, cos, sqrt, atan2
        R = 6371  # Earth radius in km
        
        lat1, lon1 = radians(coordinates[0]), radians(coordinates[1])
        lat2, lon2 = radians(center_location[0]), radians(center_location[1])
        
        dlat = lat2 - lat1
        dlon = lon2 - lon1
        
        a = sin(dlat/2)**2 + cos(lat1)*cos(lat2)*sin(dlon/2)**2
        c = 2 * atan2(sqrt(a), sqrt(1-a))
        
        return R * c
    
    def _calculate_distance_between_points(self, point1: Optional[Tuple[float, float]], point2: Optional[Tuple[float, float]]) -> float:
        """Изчислява разстоянието между две точки в км"""
        if not point1 or not point2:
            return 0.0
        
        from math import radians, sin, cos, sqrt, atan2
        R = 6371  # Earth radius in km
        
        lat1, lon1 = radians(point1[0]), radians(point1[1])
        lat2, lon2 = radians(point2[0]), radians(point2[1])
        
        dlat = lat2 - lat1
        dlon = lon2 - lon1
        
        a = sin(dlat/2)**2 + cos(lat1)*cos(lat2)*sin(dlon/2)**2
        c = 2 * atan2(sqrt(a), sqrt(1-a))
        
        return R * c
    
    def _calculate_time_between_points(self, point1: Optional[Tuple[float, float]], point2: Optional[Tuple[float, float]]) -> float:
        """Изчислява времето за пътуване между две точки в минути (приблизително)"""
        if not point1 or not point2:
            return 0.0
        
        distance_km = self._calculate_distance_between_points(point1, point2)
        # Приблизително време за пътуване: 2 минути на км (градски транспорт)
        # Това включва спирачки, светофари, задръствания и т.н.
        return distance_km * 2

    def _get_vehicle_config(self, vehicle_type):
        """Връща конфигурацията за даден тип превозно средство"""
        from config import get_config
        vehicle_configs = get_config().vehicles
        
        if vehicle_configs:
            for config in vehicle_configs:
                if config.vehicle_type == vehicle_type:
                    return config
        return None
    
    def export_warehouse_orders(self, warehouse_customers: List[Customer]) -> str:
        """Експортира заявките в склада (за съвместимост)"""
        if not warehouse_customers:
            logger.info("Няма заявки за експорт в склада")
            return ""
        
        file_path = os.path.join(self.config.excel_output_dir, self.config.warehouse_excel_file)
        os.makedirs(os.path.dirname(file_path), exist_ok=True)
        
        data = []
        for customer in warehouse_customers:
            data.append({
                'ID': customer.id,
                'Име': customer.name,
                'Обем (ст.)': customer.volume,
                'GPS координати': customer.original_gps_data,
                'Latitude': customer.coordinates[0] if customer.coordinates else '',
                'Longitude': customer.coordinates[1] if customer.coordinates else ''
            })
        
        df = pd.DataFrame(data)
        df.to_excel(file_path, index=False)
        
        logger.info(f"Складови заявки експортирани в {file_path}")
        return file_path
    
    def export_vehicle_routes(self, solution: CVRPSolution) -> str:
        """Експортира маршрутите на превозните средства (за съвместимост)"""
        file_path = os.path.join(self.config.excel_output_dir, self.config.routes_excel_file)
        os.makedirs(os.path.dirname(file_path), exist_ok=True)
        
        data = []
        for i, route in enumerate(solution.routes):
            vehicle_name = VEHICLE_SETTINGS.get(route.vehicle_type.value, {}).get('name', 'Неизвестен')
            for j, customer in enumerate(route.customers):
                data.append({
                    'Маршрут': i + 1,
                    'Превозно средство': vehicle_name,
                    'Ред в маршрута': j + 1,
                    'ID клиент': customer.id,
                    'Име клиент': customer.name,
                    'Обем (ст.)': customer.volume,
                    'GPS': customer.original_gps_data
                })
        
        df = pd.DataFrame(data)
        df.to_excel(file_path, index=False, engine='openpyxl')
        
        logger.info(f"Маршрути експортирани в {file_path}")
        return file_path


class OutputHandler:
    """Главен клас за управление на изходните данни"""
    
    def __init__(self, config: Optional[OutputConfig] = None):
        self.config = config or get_config().output
        self.excel_exporter = ExcelExporter(self.config)
    
    def generate_all_outputs(self, solution: CVRPSolution, 
                           warehouse_allocation: WarehouseAllocation,
                           depot_location: Tuple[float, float]) -> Dict[str, str]:
        """Генерира всички изходни файлове и връща речник с пътищата до тях"""
        logger.info("Започвам генериране на изходни файлове")
        output_files = {}

        # 1. Интерактивна карта
        if self.config.enable_interactive_map:
            map_gen = InteractiveMapGenerator(self.config)
            route_map = map_gen.create_map(solution, warehouse_allocation, depot_location)
            map_file = map_gen.save_map(route_map)
            output_files['map'] = map_file
        
        # 2. Обединяване на всички необслужени клиенти
        all_unserviced_customers = warehouse_allocation.warehouse_customers + solution.dropped_customers
        
        # 3. Експорт в един общ Excel файл с отделни sheets
        if solution.routes or all_unserviced_customers:
            excel_file = self.excel_exporter.export_all_to_single_excel(solution, all_unserviced_customers)
            if excel_file:
                output_files['excel_report'] = excel_file
        
        logger.info(f"Генерирани {len(output_files)} изходни файла")
        return output_files


# Удобна функция
def generate_outputs(solution: CVRPSolution, warehouse_allocation: WarehouseAllocation,
                   depot_location: Tuple[float, float]) -> Dict[str, str]:
    """Удобна функция за генериране на всички изходи"""
    handler = OutputHandler()
    return handler.generate_all_outputs(solution, warehouse_allocation, depot_location) 